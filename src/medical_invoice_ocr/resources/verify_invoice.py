# -*- coding: utf-8 -*-
"""
独立校验一份生成好的医院收据 xlsx：

- Sheet1 A1:G19 内公式数量必须为 0；
- 17 个费用值格必须是数值或空，禁止 "58-58" 之类的文本；
- C2/C3/D15/E15 必须符合 v2 模板的缩进和对齐规则；
- 若同时提供 --expected-fee-total / --expected-final，则再校验一次金额；
- 校验文件名是否为 `<image>+<name>+<YYYYMMDD>+(核对无误|核对有误).xlsx`。

用法：
    python verify-invoice.py --file <生成的 xlsx>
    python verify-invoice.py --file <xlsx> --expected-fee-total 22 --expected-final 22
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import openpyxl


FEE_CELLS: dict[str, str] = {
    "床位费": "B5",
    "诊察费": "D5",
    "西药费": "G5",
    "中成药": "B6",
    "中草药": "D6",
    "检查费": "G6",
    "治疗费": "B7",
    "放射费": "D7",
    "手术费": "G7",
    "化验费": "B8",
    "输血费": "D8",
    "输氧费": "G8",
    "放疗费": "B9",
    "护理费": "D9",
    "伙食费": "G9",
    "其他": "B10",
    "卫生材料费": "D10",
}
ALL_FEE_CELLS = list(FEE_CELLS.values())
SCAN_RANGE = "A1:G19"
FILENAME_RE = re.compile(
    r"^(?P<image>[^+\\/:*?\"<>|]+)\+(?P<name>[^+\\/:*?\"<>|]+)\+(?P<date>\d{8})\+(?P<result>核对无误|核对有误)\.xlsx$"
)


def to_cents(x) -> int:
    return int(round(float(x) * 100))


def verify(path: Path, expected_fee_total: float | None, expected_final: float | None) -> None:
    if not path.is_file():
        raise SystemExit(f"文件不存在: {path}")

    # 文件名
    fn_match = FILENAME_RE.match(path.name)
    if not fn_match:
        raise SystemExit(f"文件名不符合命名规则: {path.name}")

    wb = openpyxl.load_workbook(path)
    if "Sheet1" not in wb.sheetnames:
        raise SystemExit("Sheet1 不存在")
    ws = wb["Sheet1"]

    # v2 模板的固定缩进、对齐、经办人与开票日期。
    if not isinstance(ws["C2"].value, str) or not ws["C2"].value.startswith("    "):
        raise SystemExit(f"C2 应以 4 个空格缩进，实际: {ws['C2'].value!r}")
    if ws["C2"].alignment.horizontal != "left":
        raise SystemExit(f"C2 应左对齐，实际: {ws['C2'].alignment.horizontal!r}")
    if isinstance(ws["C3"].value, str) and ws["C3"].value != ws["C3"].value.lstrip():
        raise SystemExit(f"C3 前面不允许空格，实际: {ws['C3'].value!r}")
    if ws["C3"].alignment.horizontal != "left":
        raise SystemExit(f"C3 应左对齐，实际: {ws['C3'].alignment.horizontal!r}")
    if ws["D15"].value != "  张鸣":
        raise SystemExit(f"D15 应固定为'  张鸣'，实际: {ws['D15'].value!r}")
    if ws["D15"].alignment.horizontal != "left":
        raise SystemExit(f"D15 应左对齐，实际: {ws['D15'].alignment.horizontal!r}")
    raw_date = fn_match.group("date")
    expected_date_display = f"   {raw_date[:4]}  {raw_date[4:6]}  {raw_date[6:]}"
    if ws["E15"].value != expected_date_display:
        raise SystemExit(
            f"E15 与文件名开票日期不一致: 期望 {expected_date_display!r}，实际 {ws['E15'].value!r}"
        )

    # 页边距单位为英寸；允许 Excel 在点与英寸之间换算时产生少量舍入。
    expected_margins = {
        "left": 2.4 / 2.54,
        "right": 0.0,
        "top": 1.5 / 2.54,
        "bottom": 2.5 / 2.54,
        "header": 1.3 / 2.54,
        "footer": 1.3 / 2.54,
    }
    actual_margins = ws.page_margins
    for attr, expected in expected_margins.items():
        actual = float(getattr(actual_margins, attr))
        if abs(actual - expected) > 0.01:
            raise SystemExit(f"页边距 {attr} 不符合 v2 模板: 期望约 {expected}, 实际 {actual}")
    row11_height = ws.row_dimensions[11].height
    if row11_height is None or abs(float(row11_height) - 10.5) > 0.01:
        raise SystemExit(f"第 11 行高度应为 10.5，实际: {row11_height!r}")

    # 1. 公式扫描
    formula_cells: list[str] = []
    for row in ws[SCAN_RANGE]:
        for cell in row:
            if cell.data_type == "f":
                formula_cells.append(f"{cell.coordinate}={cell.value!r}")
    if formula_cells:
        raise SystemExit(f"发现公式（禁止）: {formula_cells}")

    # 2. 费用格：必须是数值 / int / float 或 None
    invalid_fee: list[str] = []
    invalid_fee_format: list[str] = []
    fee_values: dict[str, float] = {}
    for name, coord in FEE_CELLS.items():
        v = ws[coord].value
        dt = ws[coord].data_type
        if v is None:
            continue
        if dt == "f":
            invalid_fee.append(f"{name}({coord})=公式 {v!r}")
            continue
        if isinstance(v, bool) or not isinstance(v, (int, float)):
            invalid_fee.append(f"{name}({coord})=非数值 {v!r} data_type={dt}")
            continue
        fee_values[name] = float(v)
        if to_cents(v) == 0 and ws[coord].number_format != "0":
            invalid_fee_format.append(
                f"{name}({coord})=0 但 number_format={ws[coord].number_format!r}"
            )
    if invalid_fee:
        raise SystemExit(f"费用格违规: {invalid_fee}")
    if invalid_fee_format:
        raise SystemExit(f"零金额显示格式违规: {invalid_fee_format}")

    # 3. E12 必须是数值
    e12 = ws["E12"]
    if e12.data_type == "f":
        raise SystemExit(f"E12 是公式: {e12.value!r}")
    if not isinstance(e12.value, (int, float)):
        raise SystemExit(f"E12 不是数值: {e12.value!r} (data_type={e12.data_type})")
    final_in_sheet = float(e12.value)

    # 4. B12 必须是文本，且不以 "=" 开头
    b12 = ws["B12"]
    if b12.data_type == "f":
        raise SystemExit(f"B12 是公式: {b12.value!r}")
    if b12.value is None or not isinstance(b12.value, str):
        raise SystemExit(f"B12 应为中文大写文本，实际: {b12.value!r}")
    if b12.value.startswith("="):
        raise SystemExit(f"B12 疑似公式文本: {b12.value!r}")

    # 5. 费用格合计 vs E12
    fee_total = round(sum(fee_values.values()), 2)
    if to_cents(fee_total) != to_cents(final_in_sheet):
        # 内部一致性异常仅当文件名声称"核对无误"时抛错
        if fn_match.group("result") == "核对无误":
            raise SystemExit(
                f"文件名声称核对无误但费用格合计 {fee_total} != E12 {final_in_sheet}"
            )

    # 6. 期望值（可选）
    if expected_fee_total is not None and to_cents(fee_total) != to_cents(expected_fee_total):
        raise SystemExit(f"费用合计 {fee_total} != 期望 {expected_fee_total}")
    if expected_final is not None and to_cents(final_in_sheet) != to_cents(expected_final):
        raise SystemExit(f"E12 {final_in_sheet} != 期望 {expected_final}")

    print(f"VERIFY OK {path}")
    print(f"  filename: image={fn_match.group('image')} name={fn_match.group('name')} date={fn_match.group('date')} result={fn_match.group('result')}")
    print(f"  fee_detail_total={fee_total}  E12={final_in_sheet}  formula_count=0")
    print(f"  B12={b12.value!r}")
    print(f"  D15={ws['D15'].value!r}  E15={ws['E15'].value!r}")
    print(f"  fees={fee_values}")


def main() -> None:
    ap = argparse.ArgumentParser(description="Verify medical invoice xlsx.")
    ap.add_argument("--file", required=True)
    ap.add_argument("--expected-fee-total", type=float, default=None)
    ap.add_argument("--expected-final", type=float, default=None)
    args = ap.parse_args()

    verify(Path(args.file), args.expected_fee_total, args.expected_final)


if __name__ == "__main__":
    main()
