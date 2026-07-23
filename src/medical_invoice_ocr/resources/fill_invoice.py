# -*- coding: utf-8 -*-
"""
根据 JSON 输入将医院收据数据填入 Excel 模板并输出 xlsx。

- 每项费用 net = printed - deduction，净额 0 也写数字 0。
- E12 = invoice_totals.printed - invoice_totals.deduction，独立于费用格。
- B12 由 finalAmount 换算为中文大写普通文本。
- 命名 <原图片名称>+<交款人>+<YYYYMMDD>+<核对结果>.xlsx。
  原图片名称取 input.source_image 的文件名去扩展名后 sanitized 的结果。
- 生成后扫描 Sheet1 A1:G19，公式数量 > 0 直接报错退出。

用法:
    python fill-invoice.py --input <input.json> --template <template.xlsx> --output-dir <dir>
"""
from __future__ import annotations

import argparse
from copy import copy
import json
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl


# 费用名 → 单元格
FEE_CELLS: dict[str, str] = {
    "床位费": "B5",
    "诊察费": "D5",
    "西药费": "G5",
    "中成药": "B6",
    "中草药": "D6",
    "检查费": "G6",
    "治疗费": "B7",
    "放射费": "D7",
    "手术费": "G7",
    "化验费": "B8",
    "输血费": "D8",
    "输氧费": "G8",
    "放疗费": "B9",
    "护理费": "D9",
    "伙食费": "G9",
    "其他": "B10",
    "卫生材料费": "D10",
}

# 生成时扫描/清空的费用值格
ALL_FEE_CELLS: list[str] = list(FEE_CELLS.values())

# 有效区域（扫描公式与判定值范围）
SCAN_RANGE = "A1:G19"

# 中文大写数字
_CN_DIGITS = "零壹贰叁肆伍陆柒捌玖"
_CN_UNITS_INT = ["", "拾", "佰", "仟"]
_CN_BIG_UNITS = ["", "万", "亿"]


def _section_to_cn(n: int) -> str:
    """0-9999 -> 中文大写（未加万亿，未处理段间零）。"""
    if n == 0:
        return "零"
    digits = []
    for i in range(4):
        d = n % 10
        n //= 10
        digits.append(_CN_DIGITS[d] + _CN_UNITS_INT[i] if d else "零")
        if n == 0:
            break
    s = "".join(reversed(digits))
    while "零零" in s:
        s = s.replace("零零", "零")
    return s.rstrip("零")


def _int_to_cn(n: int) -> str:
    """正整数 -> 中文大写。"""
    if n == 0:
        return "零"
    sections: list[str] = []
    for big_i, big in enumerate(_CN_BIG_UNITS):
        if n == 0:
            break
        sec = n % 10000
        n //= 10000
        if sec == 0:
            sections.append("")
            continue
        sec_str = _section_to_cn(sec)
        # 段内不足万位补零（如 10008 -> 壹万零捌）
        if big_i > 0 and sec < 1000 and sections:
            sec_str = "零" + sec_str
        sections.append(sec_str + big)
    result = "".join(reversed(sections))
    while "零零" in result:
        result = result.replace("零零", "零")
    return result.strip("零") or "零"


def amount_to_cn_upper(amount: float) -> str:
    """
    金额转中文大写（会计写法）。
    >>> amount_to_cn_upper(22)         -> '贰拾贰元整'
    >>> amount_to_cn_upper(223)        -> '贰佰贰拾叁元整'
    >>> amount_to_cn_upper(22.5)       -> '贰拾贰元伍角整'
    >>> amount_to_cn_upper(22.05)      -> '贰拾贰元零伍分'
    >>> amount_to_cn_upper(0.5)        -> '伍角整'
    >>> amount_to_cn_upper(0)          -> '零元整'
    """
    total_fen = round(amount * 100)
    if total_fen < 0:
        return "负" + amount_to_cn_upper(-amount)
    if total_fen == 0:
        return "零元整"

    yuan = total_fen // 100
    jiao = (total_fen // 10) % 10
    fen = total_fen % 10

    yuan_part = _int_to_cn(yuan) + "元" if yuan else ""

    if jiao == 0 and fen == 0:
        return yuan_part + "整"
    if jiao == 0 and fen != 0:
        return (yuan_part + "零" if yuan else "") + _CN_DIGITS[fen] + "分"
    if fen == 0:
        return yuan_part + _CN_DIGITS[jiao] + "角整"
    return yuan_part + _CN_DIGITS[jiao] + "角" + _CN_DIGITS[fen] + "分"


def to_cents(x: float) -> int:
    """浮点转整数分。"""
    return int(round(float(x) * 100))


def _sanitize_filename_component(s: str) -> str:
    """去掉 Windows 文件名中的非法字符。"""
    return re.sub(r'[\\/:*?"<>|]', "", s).strip()


def _date_display(value: Any) -> str:
    """将 YYYYMMDD 转为模板使用的 YYYY  MM  DD；已格式化文本保持原样。"""
    raw = str(value or "").strip()
    if re.fullmatch(r"\d{8}", raw):
        return f"{raw[:4]}  {raw[4:6]}  {raw[6:]}"
    return raw


def _write_cell(ws, coord: str, value: Any) -> None:
    """写入普通值，绝不进入公式路径。"""
    cell = ws[coord]
    if isinstance(value, str) and value.startswith("="):
        # 防御性处理：强制以文本方式写入
        cell.value = None
        cell.value = "'" + value  # openpyxl 会存为字符串
        return
    cell.value = value


def load_input(path: Path) -> dict:
    raw = json.loads(path.read_text(encoding="utf-8"))
    for key in ("patient", "invoice_date", "fees", "invoice_totals", "source_image"):
        if key not in raw:
            raise ValueError(f"input JSON 缺少字段: {key}")
    if not isinstance(raw["source_image"], str) or not raw["source_image"].strip():
        raise ValueError("source_image 必须是非空字符串（原发票图片文件名或路径）")
    unknown = set(raw["fees"].keys()) - set(FEE_CELLS.keys())
    if unknown:
        raise ValueError(f"未知费用名: {sorted(unknown)}；仅支持 {list(FEE_CELLS)}")
    if not re.fullmatch(r"\d{8}", str(raw["invoice_date"])):
        raise ValueError(f"invoice_date 必须是 8 位数字 YYYYMMDD，收到 {raw['invoice_date']!r}")
    return raw


def compute_nets(fees: dict[str, dict]) -> dict[str, float]:
    nets: dict[str, float] = {}
    for name, item in fees.items():
        printed = float(item.get("printed", 0))
        deduction = float(item.get("deduction", 0))
        nets[name] = round(printed - deduction, 2)
        if nets[name] < 0:
            raise ValueError(f"费用 {name} net < 0（printed={printed}, deduction={deduction}）")
    return nets


def fill(template_path: Path, output_dir: Path, data: dict) -> dict:
    wb = openpyxl.load_workbook(template_path)
    ws = wb["Sheet1"]

    patient = data["patient"]
    _write_cell(ws, "G1", patient.get("receipt_no"))
    _write_cell(ws, "A2", _date_display(patient.get("admission_date", "")))
    discharge_date = _date_display(patient.get("discharge_date", ""))
    _write_cell(ws, "C2", f"    {discharge_date}" if discharge_date else "")
    c2_alignment = copy(ws["C2"].alignment)
    c2_alignment.horizontal = "left"
    ws["C2"].alignment = c2_alignment
    _write_cell(ws, "F2", patient.get("days"))
    _write_cell(ws, "A3", patient.get("name", ""))
    gender_raw = patient.get("gender", "")
    if gender_raw in ("男", "女"):
        gender_str = f"（{gender_raw}）"
    else:
        gender_str = str(gender_raw or "").strip()
    _write_cell(ws, "C3", gender_str)
    c3_alignment = copy(ws["C3"].alignment)
    c3_alignment.horizontal = "left"
    ws["C3"].alignment = c3_alignment

    # 先清空所有费用格
    for coord in ALL_FEE_CELLS:
        _write_cell(ws, coord, None)

    # 写入图片中出现的费用
    nets = compute_nets(data["fees"])
    fees_written: dict[str, float] = {}
    for name, net in nets.items():
        coord = FEE_CELLS[name]
        # 净额一律以数字写入，包括 0
        _write_cell(ws, coord, net)
        cell = ws[coord]
        # 统一费用显示：0 显示为“0”，非零金额显示两位小数；保持模板字体并改为红色居中。
        cell.number_format = "0" if to_cents(net) == 0 else "0.00"
        fee_font = copy(cell.font)
        fee_font.color = "FFFF0000"
        cell.font = fee_font
        fee_alignment = copy(cell.alignment)
        fee_alignment.horizontal = "center"
        cell.alignment = fee_alignment
        fees_written[name] = net

    # 计算 finalAmount（独立于费用格）
    totals = data["invoice_totals"]
    final_amount = round(float(totals["printed"]) - float(totals["deduction"]), 2)
    if final_amount < 0:
        raise ValueError(f"finalAmount < 0: printed={totals['printed']}, deduction={totals['deduction']}")

    # 计算 fee_detail_total（费用格净额之和）
    fee_detail_total = round(sum(nets.values()), 2)

    # 核对
    check_result = "核对无误" if to_cents(fee_detail_total) == to_cents(final_amount) else "核对有误"

    # B12 中文大写 / E12 数值
    _write_cell(ws, "B12", amount_to_cn_upper(final_amount))
    _write_cell(ws, "E12", final_amount)

    # 支付/固定经办人/开票日期。E15 只从用于文件命名的 invoice_date 派生，
    # 不接受识图结果中的收款员或日期显示文本，避免两处日期不一致。
    payment = data.get("payment_method", "").strip()
    _write_cell(ws, "E13", " " * 10 + payment if payment else None)
    _write_cell(ws, "D15", "  张鸣")
    d15_alignment = copy(ws["D15"].alignment)
    d15_alignment.horizontal = "left"
    ws["D15"].alignment = d15_alignment
    raw_date = str(data.get("invoice_date", ""))
    if not re.fullmatch(r"\d{8}", raw_date):
        raise ValueError(f"invoice_date 必须是 YYYYMMDD 形式的 8 位数字，实际: {raw_date!r}")
    _write_cell(ws, "E15", f"   {raw_date[:4]}  {raw_date[4:6]}  {raw_date[6:]}")
    _write_cell(ws, "E16", None)

    # 强化：E12 必须是数值类型
    e12 = ws["E12"]
    if e12.data_type == "f":
        raise RuntimeError(f"E12 竟然是公式: {e12.value!r}")
    if not isinstance(e12.value, (int, float)):
        raise RuntimeError(f"E12 不是数值: {e12.value!r} (data_type={e12.data_type})")

    # 扫描整片有效区域，禁止任何公式
    formula_count = 0
    formula_cells: list[str] = []
    for row in ws[SCAN_RANGE]:
        for cell in row:
            if cell.data_type == "f":
                formula_count += 1
                formula_cells.append(f"{cell.coordinate}={cell.value!r}")
    if formula_count > 0:
        raise RuntimeError(
            f"Sheet1 {SCAN_RANGE} 内发现 {formula_count} 个公式: {formula_cells}"
        )

    # 命名并保存
    image_stem = _sanitize_filename_component(Path(str(data["source_image"])).stem)
    if not image_stem:
        raise ValueError(f"source_image 去扩展名并 sanitize 后为空: {data['source_image']!r}")
    name = _sanitize_filename_component(patient.get("name", "未知"))
    date_str = _sanitize_filename_component(str(data["invoice_date"]))
    fname = f"{image_stem}+{name}+{date_str}+{check_result}.xlsx"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / fname
    wb.save(output_path)

    return {
        "output_path": str(output_path).replace("\\", "/"),
        "fee_detail_total": fee_detail_total,
        "final_amount": final_amount,
        "check_result": check_result,
        "formula_count": formula_count,
        "fees_written": fees_written,
    }


def main() -> None:
    ap = argparse.ArgumentParser(description="Fill medical invoice Excel template.")
    ap.add_argument("--input", required=True, help="输入 JSON 路径")
    ap.add_argument("--template", required=True, help="模板 xlsx 路径")
    ap.add_argument("--output-dir", required=True, help="输出目录")
    args = ap.parse_args()

    input_path = Path(args.input)
    template_path = Path(args.template)
    output_dir = Path(args.output_dir)

    if not input_path.is_file():
        print(f"输入不存在: {input_path}", file=sys.stderr)
        sys.exit(2)
    if not template_path.is_file():
        print(f"模板不存在: {template_path}", file=sys.stderr)
        sys.exit(2)

    data = load_input(input_path)
    try:
        result = fill(template_path, output_dir, data)
    except Exception as exc:  # noqa: BLE001
        print(f"FILL_FAILED: {exc}", file=sys.stderr)
        sys.exit(1)

    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
